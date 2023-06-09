from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import fnmatch
import os
import PyPDF2
import re
import pathlib
import shutil
from tkinter import *

root=Tk()
root.geometry('650x210')
root.title('Certificate Check')



#D:\PTM\git_pt\Pudupet\certificate
#D:\PTM\git_pt\Pudupet\Pudupet.xlsx
# Have to resign the link of certificate folder and the program have to look for the folder with udise name and school name  
'''loc_cert='D:\\PTM\\git_pt\\Pudupet\\certificate'

# Have to resing the link the excel with the student data 
loc_excel='D:\\PTM\\git_pt\\Pudupet\\Pudupet.xlsx'''

def check_data():

	loc_excel=loca_excel.get()
	loc_cert=loca_cert.get()
	#print(loc_excel,loc_cert)
	wb=load_workbook(loc_excel)
	ws=wb.active

	# New workbook for checked details
	new=Workbook()
	checked=new.active
	absent=new.create_sheet('Absent')
	checked['A1']='S.No.'
	checked['B1']='Student Name'
	checked['C1']='Father Name'
	checked['D1']='Aadhar No.'
	checked['E1']='Roll No.'

	absent['A1']='S.No.'
	absent['B1']='District'
	absent['C1']='School Name'
	absent['D1']='UDISE'
	absent['E1']='Student Name'
	absent['F1']='Father Name'
	absent['G1']='Aadhar No.'
	absent['H1']='Roll No.'

	s_no=1
	data=[]
	a=ws['L2'].value
	row=2
	a_list=[]
	p_list=[]
	rl=[]

	# To get the total number of active rows 
	while a is not None:
		a=ws['L'+str(row)].value
		row=row+1
		rl.append(a)

	# Getting the students details
	for i in range(2,row-1):
		s={}
		s['s_name']=ws['G'+str(i)].value.title()
		s['s_fa_name']=ws['O'+str(i)].value.title()
		s['s_aadhar']=ws['K'+str(i)].value
		s['s_roll']=ws['L'+str(i)].value
		s['district']=ws['B'+str(i)].value
		s['school']=ws['C'+str(i)].value
		s['udise']=ws['D'+str(i)].value
		s['emis']=ws['N'+str(i)].value
		data.append(s)

	fill=PatternFill(patternType='solid',
						fgColor='00FF0000')

	# Checking the details of students with the certificate
	source =os.listdir(loc_cert)
	for fol in source:
		loc= loc_cert+'\\'+fol
		name=fnmatch.filter(os.listdir(loc),'*_Tamil Nadu School Certificate_*.pdf')
		for f_name in name:
			file=open(loc+'\\'+f_name,'rb')
			reader=PyPDF2.PdfReader(file)
			for page in reader.pages:
				text=page.extract_text()
				pattern=re.compile(r'[0-9]{7}')
				match=pattern.findall(text)
				roll=match[0]
				for i in range(0,len(data)):
					if roll in data[i]['s_roll']:

						p_name=re.compile(data[i]['s_name'])
						p_fa_name=re.compile(data[i]['s_fa_name'])
						p_aadhar=re.compile('XXXXXXXX'+r'[0-9]{4}')
						m_name=p_name.findall(text)
						m_fa_name=p_fa_name.findall(text)
						m_aadhar=p_aadhar.findall(text)
						aadhar_c=[str(y) for y in m_aadhar[0]]
						aadhar_f=[str(x) for x in data[i]['s_aadhar']]
						#print(m_name,m_fa_name,m_aadhar)
						checked['A'+str(s_no+1)]=s_no
						s_no += 1
						if len(m_name) > 0:
							checked['B'+str(s_no)]=m_name[0] 
						else:
							checked['B'+str(s_no)]=data[i]['s_name'] 
							checked['B'+str(s_no)].fill=fill
						try:
							if m_fa_name != None:
								checked['C'+str(s_no)]=m_fa_name[0]
							else:
								checked['C'+str(s_no)]=data[i]['s_fa_name']
								checked['B'+str(s_no)].fill=fill
						except:
							checked['C'+str(s_no)]=data[i]['s_fa_name']
							checked['B'+str(s_no)].fill=fill

						try:
							if aadhar_c[8]==aadhar_f[8] and aadhar_c[9]==aadhar_f[9] and aadhar_c[10]==aadhar_f[10] and  aadhar_c[11]==aadhar_f[11]:
								checked['D'+str(s_no)]=data[i]['s_aadhar']
							else:
								checked['D'+str(s_no)]=data[i]['s_aadhar']
								checked['B'+str(s_no)].fill=fill
						except:
							checked['D'+str(s_no)]=data[i]['s_aadhar']
							checked['B'+str(s_no)].fill=fill

						checked['E'+str(s_no)]=data[i]['s_roll']
						p_list.append(roll)
						
			file.close()

	s=1
	for i in range(0,len(data)):
		if data[i]['s_roll'] not in p_list:
			absent['A'+str(s+1)]=s
			absent['B'+str(s+1)]=data[i]['district']
			absent['C'+str(s+1)]=data[i]['school']
			absent['D'+str(s+1)]=data[i]['udise']
			absent['E'+str(s+1)]=data[i]['s_name']
			absent['F'+str(s+1)]=data[i]['s_fa_name']
			absent['G'+str(s+1)]=data[i]['s_aadhar']
			absent['H'+str(s+1)]=data[i]['s_roll']
			s+=1


	new.save('Checked.xlsx')
	fin=Label(root,text='Data Checked and stored successfully')
	fin.grid(row=5, column=0, columnspan=2)


loca_cert_lab=Label(root, text='Enter the location of Certificate folder:')
loca_cert_lab.grid(row=0,column=0, pady=10)
loca_cert=Entry(root, width=60)
loca_cert.grid(row=0,column=1)
loca_cert.insert(0,"c:\\User\\Nk\\Desktop")

loca_excel_lab=Label(root, text='Enter the location of excel:')
loca_excel_lab.grid(row=1,column=0, pady=20)
loca_excel=Entry(root,width=60)
loca_excel.grid(row=1,column=1)
loca_excel.insert(0,'c:\\User\\Nk\\Desktop\\workbook.xlsx')


check=Button(root, text='Check', command=check_data)
check.grid(row=2, column=0, columnspan=2)

message=Label(root,text='After Click Check\n'
                            'Dont close or click the check button again, until success message popup')
message.grid(row=3, column=0, columnspan=2)

message=Label(root,text='==============================================================================')
message.grid(row=4, column=0, columnspan=2)

root.mainloop()